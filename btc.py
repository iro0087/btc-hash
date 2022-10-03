#do not find the last version of my code but you got the idea (renaming the xlsx file, why not adapt the code for eth,
#and find the correct class to grab)

pu2 = 3212

import os

import time

from playsound import playsound

import sys

while True:

    import requests

    from bs4 import BeautifulSoup

    from openpyxl import load_workbook

    type_l = ["btc", "btc"]

    start = time.time()

    nb = 0

    tf = 0

    stop = 0

    while tf < len(type_l):

        pu2 = int(pu2)

        pu2 = pu2 + 1

        if pu2 == 13500:

            playsound("mus.wav")

            sys.exit()

        pu2 = str(pu2)

        workbook2 = load_workbook(filename="cryp" + type_l[tf] + "2.xlsx")

        sheet2 = workbook2.active

        tb = sheet2.cell(row=1, column=1)

        t2b = sheet2.cell(row=1, column=2)

        t3b = sheet2.cell(row=1, column=3)

        workbook = load_workbook(filename="puter.xlsx")

        sheet = workbook.active

        t = tb.value

        print(t)

        t = int(t)

        t2 = t2b.value

        t2 = int(t2)

        print("hgfdf", pu2)

        url = requests.get("https://www.blockchain.com/" + type_l[tf] + "/blocks?page=" + pu2)  

        scraping = BeautifulSoup(url.text, "html.parser")

        if url.ok:

            for titre in scraping.find_all(attrs={"class": "rjh6gp-0"}):  #may cause error

                count_zero = 0

        titre2 = str(titre)

        titre2b = " ".join(titre2)

        l_all = str.split(titre2b)

        l_allf = []

        l_allf2 = []

        l_all3 = []

        t_hsh = 0

        while t_hsh < len(l_all):

            if l_all[t_hsh] == "0" and l_all[t_hsh + 1] == "0":

                count_zero = count_zero + 1

            if count_zero > 2:

                l_allf.insert(len(l_allf) + 1, l_all[t_hsh])

            if t_hsh + 9 < len(l_all):

                if l_all[t_hsh + 9] == "=":

                    count_zero = 0

            t_hsh = t_hsh + 1

        t_hsh = 0

        t_hsh2 = 0

        l_all = []

        while t_hsh < len(l_allf) and stop == 0:

            l_all.insert(t_hsh, l_allf[t_hsh])

            if t_hsh2 == 61 or t_hsh + 1 == len(l_allf):

                l_all.insert(0, "0")

                l_all.insert(0, "0")

                l_all_str = "".join(l_all)

                if sheet2.cell(row=1, column=4).value != None:

                    if sheet2.cell(row=1, column=4).value == l_all_str:

                        stop = 1

                    else:

                        l_all3.insert(t_hsh, l_all_str)

                        t_hsh2 = -1

                        l_all = []

                else:

                    l_all3.insert(t_hsh, l_all_str)

                    t_hsh2 = -1

                    l_all = []

            t_hsh2 = t_hsh2 + 1

            t_hsh = t_hsh + 1

        t3 = t3b.value + len(l_all3)

        t3 = int(t3)

        t4 = 0

        stop = 0

        t_hsh = 0

        while t4 < len(l_all3):

            if l_all3[t4] == l_all3[t4 - 1]:

                l_all3.remove(l_all3[t4 - 1])

            t4 = t4 + 1

        t4 = 0

        while t4 < len(l_all3):

            sheet.cell(row=t2, column=t).value = l_all3[t4]

            t = t + 1

            t4 = t4 + 1

            if t + t3 > 1000:

                t = 1

                t2 = t2 + 1

                t3 = 1

        t4 = 0

        if len(l_all3) > 1:

            sheet2.cell(row=1, column=4).value = l_all3[0]

        sheet2.cell(row=1, column=3).value = t3

        workbook.save(filename="Cryp" + type_l[tf] + ".xlsx")

        sheet2.cell(row=1, column=1).value = t

        sheet2.cell(row=1, column=2).value = t2

        workbook2.save(filename="cryp" + type_l[tf] + "2.xlsx")

        end = time.time()

        print("réalisé en:", end - start, "s")

        tf = tf + 1

    tf = 0





























